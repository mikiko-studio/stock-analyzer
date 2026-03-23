"""
utils/ui_helpers.py
Shared formatting and UI helper functions for all pages.
"""

import io
from datetime import datetime

import pandas as pd
import streamlit as st


def format_pct(val, decimals=2):
    """Format a float as percentage string, or '—' if None/NaN."""
    if val is None:
        return "—"
    if isinstance(val, float) and pd.isna(val):
        return "—"
    try:
        return f"{val:.{decimals}%}"
    except (TypeError, ValueError):
        return "—"


def format_currency(val, currency="JPY"):
    """Format price with currency symbol."""
    if val is None:
        return "—"
    try:
        if pd.isna(val):
            return "—"
    except (TypeError, ValueError):
        pass
    sym = "¥" if currency == "JPY" else "$"
    try:
        if currency == "JPY":
            return f"{sym}{val:,.0f}"
        return f"{sym}{val:,.2f}"
    except (TypeError, ValueError):
        return "—"


def format_number(val, decimals=2):
    """Format a float as number string, or '—' if None/NaN."""
    if val is None:
        return "—"
    try:
        if pd.isna(val):
            return "—"
        return f"{val:.{decimals}f}"
    except (TypeError, ValueError):
        return "—"


def signal_badge(label, color):
    """Return an HTML badge for status indicators."""
    return (
        f'<span style="background:{color};color:#fff;padding:2px 8px;'
        f'border-radius:4px;font-size:0.8rem;font-weight:bold;">{label}</span>'
    )


def score_color(score, max_score=2):
    """Map a score (-2 to +2) to a color."""
    if score >= max_score:
        return "#22c55e"
    if score >= 1:
        return "#86efac"
    if score >= 0:
        return "#fbbf24"
    if score >= -1:
        return "#fb923c"
    return "#ef4444"


def render_metric_card(label, value, delta=None, help_text=None):
    """Render a styled metric card using st.metric."""
    st.metric(label=label, value=value, delta=delta, help=help_text)


def hero_header(title, subtitle, icon="📊"):
    """Render a hero header section."""
    st.markdown(f"# {icon} {title}")
    st.markdown(f"**{subtitle}**")
    st.divider()


def status_badge_html(status):
    """Return colored HTML badge for pass/fail/skip status."""
    color_map = {
        "passed": "#22c55e",
        "failed": "#ef4444",
        "skipped": "#94a3b8",
        "error": "#f97316",
    }
    label_map = {
        "passed": "✅ 合格",
        "failed": "❌ 不合格",
        "skipped": "⏭️ スキップ",
        "error": "⚠️ エラー",
    }
    color = color_map.get(status, "#94a3b8")
    label = label_map.get(status, status)
    return signal_badge(label, color)


def make_color_scale_df(df, col, reverse=False):
    """Apply background color gradient to a DataFrame column."""
    return df.style.background_gradient(subset=[col], cmap="RdYlGn" if not reverse else "RdYlGn_r")


def _to_excel_bytes(df: pd.DataFrame, pct_cols: list[str] | None = None,
                    float2_cols: list[str] | None = None) -> bytes:
    """
    Convert DataFrame to Excel bytes with styled header and number formatting.

    pct_cols  : columns whose values are 0.0–1.0 fractions → formatted as "0.00%"
    float2_cols: columns with plain floats → formatted as "#,##0.00"
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    pct_cols = pct_cols or []
    float2_cols = float2_cols or []

    wb = Workbook()
    ws = wb.active
    ws.title = "結果"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")

    # ── Header row ───────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws.row_dimensions[1].height = 20

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, (col_name, value) in enumerate(zip(df.columns, row), 1):
            # Convert pandas NA/NaN to None so openpyxl writes blank
            if value is not None:
                try:
                    import math
                    if isinstance(value, float) and math.isnan(value):
                        value = None
                except (TypeError, ValueError):
                    pass
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in pct_cols and isinstance(value, (int, float)):
                cell.number_format = "0.00%"
            elif col_name in float2_cols and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"

    # ── Auto column width ─────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for val in df[col_name]:
            if val is not None:
                try:
                    import math
                    if isinstance(val, float) and math.isnan(val):
                        continue
                except (TypeError, ValueError):
                    pass
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_export_buttons(df: pd.DataFrame, filename_prefix: str,
                          pct_cols: list[str] | None = None,
                          float2_cols: list[str] | None = None) -> None:
    """
    Render side-by-side CSV and Excel download buttons below a results table.

    filename_prefix : e.g. "dividend_screener" → "dividend_screener_20260322_1430.csv"
    pct_cols        : column names whose values are fractions (0.0–1.0) for % Excel format
    float2_cols     : column names with plain floats for "#,##0.00" Excel format
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base = f"{filename_prefix}_{timestamp}"

    csv_bytes = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
    xlsx_bytes = _to_excel_bytes(df, pct_cols=pct_cols, float2_cols=float2_cols)

    btn_col1, btn_col2, _ = st.columns([1, 1, 2])
    with btn_col1:
        st.download_button(
            label="📥 CSV ダウンロード",
            data=csv_bytes,
            file_name=f"{base}.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with btn_col2:
        st.download_button(
            label="📊 Excel ダウンロード",
            data=xlsx_bytes,
            file_name=f"{base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
